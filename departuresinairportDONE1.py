import requests
import pandas as pd
import json
import time
from datetime import datetime
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os


def fetch_departures():
    url = "https://www.cial.aero/Flights/Departures"
    response = requests.get(url)
    return response.content


def extract_flight_data(html_content):
    soup = BeautifulSoup(html_content, 'lxml')
    table_rows = soup.find_all("div", class_="row chart-bg")
    flights = []
    for row in table_rows:
        flight_data = {}
        columns = row.find_all("span")
        if len(columns) > 0:
            flight_data["Airline"] = columns[1].text.strip()
            flight_data["Origin/Destination"] = columns[3].text.strip()
            flight_data["Date/Status"] = columns[5].text.strip()
            original_sch = columns[7].text.strip()
            original_actual_time = columns[9].text.strip()
            flight_data["SCH"] = reformat_date_time(original_sch)
            flight_data["Actual Time"] = reformat_date_time(original_actual_time)

            # Check if the flight is delayed or on-time
            flight_data["Flight Status"] = "ON-TIME" if flight_data["Actual Time"] <= flight_data["SCH"] else "DELAYED"
            flight_data["Terminal"] = columns[11].text.strip()
            flights.append(flight_data)

    return flights


def reformat_date_time(date_str):
    try:
        dt = datetime.strptime(date_str, "%d %b %Y %H:%M")
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    except ValueError:
        return date_str


# Function to save data to JSON file
def save_to_json(data):
    filename = "departures.json"
    with open(filename, "w") as json_file:
        json.dump(data, json_file, indent=4)
        print(f"Flight data saved to {filename}")

# Function to save data to Excel
def save_to_excel(flight_data, timestamp):
    df = pd.DataFrame(flight_data)
    df['Timestamp'] = timestamp
    excel_filename = "Departures.xlsx"

    if not os.path.exists(excel_filename):
        # Create a new workbook if the file does not exist
        df.to_excel(excel_filename, index=False, sheet_name='Departures', header=True)
        print(f"New Excel file created: {excel_filename}")
    else:
        with pd.ExcelWriter(excel_filename, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
            start_row = writer.sheets['Departures'].max_row  # Get the last row
            df.to_excel(writer, index=False, sheet_name='Departures', startrow=start_row, header=False)
            print(f"Flight data appended to {excel_filename}")

    # Apply color coding
    try:
        wb = load_workbook(excel_filename)
        ws = wb['Departures']
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            flight_status_cell = row[df.columns.get_loc("Flight Status")]
            flight_status = flight_status_cell.value
            if flight_status == "ON-TIME":
                flight_status_cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            elif flight_status == "DELAYED":
                flight_status_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        wb.save(excel_filename)
        print(f"Color coding applied to {excel_filename}")
    except Exception as e:
        print(f"An error occurred during color coding: {e}")

if __name__ == '__main__':
    departures = fetch_departures()
    if departures:
        flight_data = extract_flight_data(departures)
        if flight_data:
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            data_with_timestamp = {"timestamp": timestamp, "flights": flight_data}
            print(json.dumps(data_with_timestamp, indent=4))
            save_to_json(data_with_timestamp)
            save_to_excel(flight_data, timestamp)
time.sleep(120)
